import openpyxl
from openpyxl import load_workbook
import os
import time
import sys
from datetime import datetime
import os
import time
from openpyxl.styles import PatternFill
from openpyxl.styles import Color, Font, Alignment



import logging
import logging.config
CON_LOG='log.conf'
logging.config.fileConfig(CON_LOG)
logging=logging.getLogger()

Excel_Headers={
    "isrun":1,
    "sort":2,
    "id":3,
    "name":4,
    "engine":5,
    "nlpVersion":6,
    "query":7,
    "retCode":8,
    "category":9,
    "domain":10,
    "action":11,
    "isDialog":12,
    "expect_response":13,
    "expect_results":14,
    "sync_tts":15,
    "IsPass":16,
    "failReason":17,
    "real_result":18
}

logging.info("Excel表头")
logging.info(Excel_Headers)

def handle_casedata(filepath,sheet=None):
    wb=load_workbook(filepath)
    if sheet is None:
        ws=wb.active
        #print(ws)
    else:
        ws=wb[sheet]
        #print(ws)

    row_max=ws.max_row
    con_max=ws.max_column


    con_titles=[]


    #提取用例
    Cases=[]
    logging.info("提取用例")
    for j in range(1,row_max+1):
        tmp_case={}
        for head in Excel_Headers:
            tmp_case[head]=ws.cell(j+1,Excel_Headers[head]).value
        if tmp_case['id']==None:
            #判断id是否是空
            break
        Cases.append(tmp_case)
    #logging.info("获取的用例： ")
    logging.info(Cases)
    return Cases


def creat_excel(filename=None):
    dt = datetime.now()
    now_time = dt.strftime('%Y_%m_%d_%H_%M_%S')  # 得用下划线，用： 号无法截图保存
    my_path = os.path.abspath(os.getcwd())
    if filename == None:
        filename = my_path + '/Result/result_%s.xlsx' % (now_time)
    else:  # 为了gui加上的，ui页面中要选中路径，传进来的是路径，不包括文件名
        filename = filename + '/result_%s.xlsx' % (now_time)

    wb = openpyxl.Workbook()
    mysheet = wb.active

    mysheet.merge_cells('A1:R1')
    mysheet.cell(row=1, column=1, value="NLP接口测试结果")
    mysheet.row_dimensions[1].height = 25

    # 然后如下设置：
    # 设置表头字体居中
    mycell = mysheet['A1']
    mycell.font = Font(name=u'宋体', bold=True)
    mycell.alignment = Alignment(horizontal='center', vertical='center')

    #result_head = ['次数', '结果', '原始日志路径', '截取后日志路径', 'audio文件路径', '设备ID']
    for i, item in enumerate(Excel_Headers):
        print(i, item)
        mysheet.cell(row=2, column=i + 1, value=item).alignment = Alignment(horizontal='center', vertical='center')
    # mysheet['F2'].font=Font(name=u'宋体',bold=True)
    # mysheet['H2'].font=Font(name=u'宋体',bold=True)

    # mysheet.column_dimensions['A'].width = 40
    # mysheet.column_dimensions['B'].width = 40
    # mysheet.column_dimensions['C'].width = 60
    # mysheet.column_dimensions['D'].width = 60
    # mysheet.column_dimensions['E'].width = 60
    # mysheet.column_dimensions['F'].width = 60

    mysheet.title = "测试结果"
    # mysheet.row_dimensions[3].height=25  #设置行高,设置第3行的行高

    wb.save(filename)
    # print(filename)
    logging.info("结果excel保存路径： " + str(filename))
    return filename

def copy_excel(filepath,respath=None):
    pass

def save_excel(excelpath,case,flag=None,fail_reasons=None,response=None,sign=None,sheet=None):
    wb = load_workbook(excelpath)
    if sheet is None:
        ws = wb.active
        # print(ws)
    else:
        ws = wb[sheet]
        # print(ws)
    result=case
    result['IsPass']=flag
    result['failReason']=fail_reasons
    result['real_result']=response
    row_max=ws.max_row
    i=1
    # print(result)00FF00
    fill_fail = openpyxl.styles.PatternFill("solid",fgColor="FF0000")
    fill_pass=openpyxl.styles.PatternFill("solid",fgColor="00FF00")
    sign_1=openpyxl.styles.PatternFill("solid",fgColor="FFFF40")
    for k, v in result.items():

        if v==[]:
            v=' '
        if v==None:
            v=''
        ws.cell(row=row_max+1, column=i).value=str(v)
        if str(v)=="False":
            ws.cell(row=row_max+1, column=i).fill=fill_fail
        elif str(v)=="True":
            ws.cell(row=row_max + 1, column=i).fill = fill_pass
            if sign==1:
                ws.cell(row=row_max + 1, column=i).fill = sign_1
            elif sign==2:
                ws.cell(row=row_max + 1, column=i).fill = sign_1
            elif sign==3:
                ws.cell(row=row_max + 1, column=i).fill = sign_1
            elif sign==4:
                ws.cell(row=row_max + 1, column=i).fill = sign_1
            elif sign==5:
                ws.cell(row=row_max + 1, column=i).fill = sign_1
            elif sign==6:
                ws.cell(row=row_max + 1, column=i).fill = sign_1
            # elif sign==7:
            #     ws.cell(row=row_max + 1, column=i).fill = sign_1
            else:
                ws.cell(row=row_max + 1, column=i).fill = fill_pass
        # elif sign==1:  #设备离线
        #     ws.cell(row=row_max + 1, column=i).fill = sign_1
        i=i+1
    wb.save(excelpath)






if __name__=="__main__":
    # filepath="D:\\Python_Project\\yuyin_request\\Nlp_interface\\case\\case.xlsx"
    # Cases=handle_casedata(filepath,'Sheet1')
    # print("========")
    # print(Cases)

    creat_excel()